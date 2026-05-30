import requests
import openpyxl
import smtplib
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime

# ============================================================
# LEER CONFIGURACION
# ============================================================

def leer_config():
    config = {}
    ruta_config = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
    with open(ruta_config, "r", encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if linea and not linea.startswith("#"):
                clave, valor = linea.split("=", 1)
                config[clave.strip()] = valor.strip()
    return config

def normalizar(texto):
    return " ".join(str(texto).split()).upper()

# ============================================================
# PASO 1: BAJAR STOCK DESDE BSALE
# ============================================================

def obtener_stock_bsale(token):
    print("Bajando stock desde Bsale...")
    headers = {"access_token": token}
    stock = {}
    offset = 0
    limit = 50

    while True:
        url = f"https://api.bsale.cl/v1/stocks.json?limit={limit}&offset={offset}&expand=[variant,office]"
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print(f"Error Bsale: {response.status_code}")
            break
        data = response.json()
        items = data.get("items", [])
        if not items:
            break
        for item in items:
            try:
                variant  = item.get("variant", {})
                office   = item.get("office", {})
                codigo   = normalizar(variant.get("code", ""))
                sucursal = normalizar(office.get("name", ""))
                nombre   = str(variant.get("description", "")).strip()
                cantidad = float(item.get("quantityAvailable", 0) or 0)
                if codigo:
                    key = (codigo, sucursal)
                    if key not in stock:
                        stock[key] = {"cantidad": 0, "nombre": nombre}
                    stock[key]["cantidad"] += cantidad
            except:
                continue
        offset += limit
        if offset >= data.get("count", 0):
            break

    print(f"Stock bajado: {len(stock)} registros")
    return stock

# ============================================================
# PASO 2: LLENAR HOJA ENTRADA
# ============================================================

def actualizar_excel(ruta_excel, stock_bsale):
    print("Actualizando hoja ENTRADA...")
    wb = openpyxl.load_workbook(ruta_excel)

    if "ENTRADA" not in wb.sheetnames:
        print("ERROR: No existe la hoja ENTRADA")
        return False

    ws = wb["ENTRADA"]

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    ws.cell(row=1, column=1).value = "PRODUCTOS"
    ws.cell(row=1, column=2).value = "TRUB"
    ws.cell(row=1, column=3).value = "SUCURSAL"
    ws.cell(row=1, column=4).value = "STOCK"

    fila = 2
    for (codigo, sucursal), datos in stock_bsale.items():
        ws.cell(row=fila, column=1).value = datos["nombre"]
        ws.cell(row=fila, column=2).value = codigo
        ws.cell(row=fila, column=3).value = sucursal
        ws.cell(row=fila, column=4).value = int(datos["cantidad"])
        fila += 1

    wb.save(ruta_excel)
    print(f"ENTRADA actualizada con {fila - 2} filas")
    return True

# ============================================================
# PASO 3: LEER HOJA PRODUCCION
# ============================================================

def leer_produccion(ruta_excel, stock_bsale):
    print("Leyendo hoja PRODUCCION...")
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb["PRODUCCION"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[normalizar(val)] = col

    col_cm       = headers.get("CM")
    col_producto = headers.get("PRODUCTOS") or headers.get("PRODUCTO")
    col_promedio = headers.get("PROMEDIO X MES") or headers.get("PROMEDIO")
    col_cocinero = headers.get("COCINERO")

    sucursales = list(set(s for (_, s) in stock_bsale.keys()))

    productos = []
    for row in range(2, ws.max_row + 1):
        cm       = ws.cell(row=row, column=col_cm).value if col_cm else None
        producto = ws.cell(row=row, column=col_producto).value if col_producto else None
        promedio = ws.cell(row=row, column=col_promedio).value if col_promedio else 0
        cocinero = ws.cell(row=row, column=col_cocinero).value if col_cocinero else ""

        if not producto or str(producto).strip() == "":
            continue

        codigo = normalizar(cm) if cm else ""

        stock_por_sucursal = {}
        total = 0
        for sucursal in sucursales:
            key = (codigo, sucursal)
            cantidad = int(stock_bsale.get(key, {}).get("cantidad", 0))
            stock_por_sucursal[sucursal] = cantidad
            total += cantidad

        vitacura = stock_por_sucursal.get("VITACURA", 0)
        pataguas = stock_por_sucursal.get("LAS PATAGUAS", 0)

        try:
            prom = float(promedio or 0)
        except:
            prom = 0

        if prom > 0:
            dias = round(total / (prom / 30), 1)
        else:
            dias = None

        productos.append({
            "producto":  str(producto).strip(),
            "total":     total,
            "vitacura":  vitacura,
            "pataguas":  pataguas,
            "promedio":  prom,
            "dias":      dias,
            "cocinero":  str(cocinero).strip() if cocinero else "-",
        })

    # Ordenar por dias restantes (None al final)
    productos.sort(key=lambda p: (p["dias"] is None, p["dias"] if p["dias"] is not None else 9999))
    return productos

# ============================================================
# PASO 4: ARMAR CORREO HTML
# ============================================================

def dias_label(dias):
    if dias is None:
        return "-"
    if dias == 0:
        return "0 días"
    return f"{dias} días"

def armar_html(productos):
    fecha = datetime.now().strftime("%A %d de %B de %Y · %H:%M")

    sin_stock = [p for p in productos if p["total"] == 0]
    criticos  = [p for p in productos if p["total"] > 0 and (p["dias"] is not None and p["dias"] <= 2)]
    bajos     = [p for p in productos if p["total"] > 0 and (p["dias"] is None or p["dias"] > 2) and p["total"] <= 8]
    ok        = [p for p in productos if p["total"] > 8 and (p["dias"] is None or p["dias"] > 2)]

    def badge(tipo):
        estilos = {
            "sin_stock": "background:#FCEBEB;color:#A32D2D;",
            "critico":   "background:#FAEEDA;color:#854F0B;",
            "bajo":      "background:#FAF3DA;color:#7A6010;",
        }
        textos = {
            "sin_stock": "Sin stock",
            "critico":   "Crítico",
            "bajo":      "Bajo stock",
        }
        return f'<span style="display:inline-block;font-size:11px;font-weight:500;padding:2px 8px;border-radius:99px;{estilos[tipo]}">{textos[tipo]}</span>'

    def color_dias(dias):
        if dias is None:
            return "#888780"
        if dias <= 1:
            return "#A32D2D"
        if dias <= 3:
            return "#854F0B"
        if dias <= 7:
            return "#3B6D11"
        return "#3B6D11"

    def filas_tabla(lista, tipo):
        html = ""
        for p in lista:
            cd = color_dias(p["dias"])
            b  = badge(tipo) + "&nbsp; " if tipo != "ok" else ""
            html += f"""
        <tr>
          <td style="padding:8px 10px;font-size:13px;border-bottom:0.5px solid #e5e5e2;">{b}{p['producto']}</td>
          <td style="padding:8px 10px;font-size:13px;text-align:right;border-bottom:0.5px solid #e5e5e2;">{p['vitacura']}</td>
          <td style="padding:8px 10px;font-size:13px;text-align:right;border-bottom:0.5px solid #e5e5e2;">{p['pataguas']}</td>
          <td style="padding:8px 10px;font-size:13px;text-align:right;font-weight:500;border-bottom:0.5px solid #e5e5e2;">{p['total']}</td>
          <td style="padding:8px 10px;font-size:13px;text-align:right;color:{cd};font-weight:500;border-bottom:0.5px solid #e5e5e2;">{dias_label(p['dias'])}</td>
          <td style="padding:8px 10px;font-size:13px;border-bottom:0.5px solid #e5e5e2;">{p['cocinero']}</td>
        </tr>"""
        return html

    def tabla(titulo, lista, tipo, color_titulo):
        if not lista:
            return ""
        return f"""
    <p style="font-size:12px;font-weight:500;text-transform:uppercase;letter-spacing:0.05em;color:{color_titulo};margin:20px 0 6px;">{titulo}</p>
    <table style="width:100%;border-collapse:collapse;border:0.5px solid #e5e5e2;border-radius:8px;overflow:hidden;margin-bottom:4px;">
      <thead>
        <tr style="background:#f5f4f0;">
          <th style="padding:8px 10px;font-size:11px;font-weight:500;color:#888780;text-align:left;border-bottom:0.5px solid #e5e5e2;">Producto</th>
          <th style="padding:8px 10px;font-size:11px;font-weight:500;color:#888780;text-align:right;border-bottom:0.5px solid #e5e5e2;">Vitacura</th>
          <th style="padding:8px 10px;font-size:11px;font-weight:500;color:#888780;text-align:right;border-bottom:0.5px solid #e5e5e2;">Pataguas</th>
          <th style="padding:8px 10px;font-size:11px;font-weight:500;color:#888780;text-align:right;border-bottom:0.5px solid #e5e5e2;">Total</th>
          <th style="padding:8px 10px;font-size:11px;font-weight:500;color:#888780;text-align:right;border-bottom:0.5px solid #e5e5e2;">Días</th>
          <th style="padding:8px 10px;font-size:11px;font-weight:500;color:#888780;text-align:left;border-bottom:0.5px solid #e5e5e2;">Cocinero</th>
        </tr>
      </thead>
      <tbody>{filas_tabla(lista, tipo)}</tbody>
    </table>"""

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:20px;background:#f5f4f0;font-family:Arial,sans-serif;">
  <div style="max-width:640px;margin:0 auto;">

    <div style="background:#ffffff;border:0.5px solid #e5e5e2;border-radius:12px;padding:20px 24px;margin-bottom:12px;">
      <p style="font-size:18px;font-weight:500;margin:0 0 4px;color:#2c2c2a;">Reporte de producción</p>
      <p style="font-size:13px;color:#888780;margin:0;">{fecha} &nbsp;·&nbsp; Vitacura y Pataguas</p>
    </div>

    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px;">
      <div style="background:#FCEBEB;border-radius:8px;padding:12px;text-align:center;">
        <div style="font-size:24px;font-weight:500;color:#A32D2D;">{len(sin_stock)}</div>
        <div style="font-size:11px;color:#A32D2D;margin-top:2px;">Sin stock</div>
      </div>
      <div style="background:#FAEEDA;border-radius:8px;padding:12px;text-align:center;">
        <div style="font-size:24px;font-weight:500;color:#854F0B;">{len(criticos)}</div>
        <div style="font-size:11px;color:#854F0B;margin-top:2px;">Crítico</div>
      </div>
      <div style="background:#EAF3DE;border-radius:8px;padding:12px;text-align:center;">
        <div style="font-size:24px;font-weight:500;color:#3B6D11;">{len(bajos)}</div>
        <div style="font-size:11px;color:#3B6D11;margin-top:2px;">Bajo stock</div>
      </div>
      <div style="background:#E1F5EE;border-radius:8px;padding:12px;text-align:center;">
        <div style="font-size:24px;font-weight:500;color:#0F6E56;">{len(ok)}</div>
        <div style="font-size:11px;color:#0F6E56;margin-top:2px;">OK</div>
      </div>
    </div>

    <div style="background:#ffffff;border:0.5px solid #e5e5e2;border-radius:12px;padding:16px 20px;">
      {tabla("Sin stock y crítico — producir hoy", sin_stock + criticos, "sin_stock", "#A32D2D")}
      {tabla("Bajo stock — producir esta semana", bajos, "bajo", "#854F0B")}
      {tabla("OK — sin urgencia", ok, "ok", "#3B6D11")}
    </div>

    <p style="font-size:12px;color:#888780;text-align:center;margin-top:12px;">Generado automáticamente · Stock actualizado desde Bsale</p>
  </div>
</body>
</html>"""
    return html

# ============================================================
# PASO 5: ENVIAR CORREO
# ============================================================

def enviar_correo(config, html):
    print("Enviando correo...")
    remitente     = config["GMAIL_REMITENTE"]
    password      = config["GMAIL_PASSWORD"]
    destinatarios = [c.strip() for c in config["CORREOS_DESTINO"].split(",")]
    fecha         = datetime.now().strftime("%d/%m/%Y")

    msg = MIMEMultipart("alternative")
    msg["From"]    = remitente
    msg["To"]      = ", ".join(destinatarios)
    msg["Subject"] = f"Reporte de Produccion - {fecha}"
    msg.attach(MIMEText(html, "html", "utf-8"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(remitente, password)
        server.sendmail(remitente, destinatarios, msg.as_string())

    print(f"Correo enviado a: {', '.join(destinatarios)}")

# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 50)
    print("INICIANDO REPORTE DE STOCK")
    print("=" * 50)

    config      = leer_config()
    stock_bsale = obtener_stock_bsale(config["BSALE_TOKEN"])

    if not actualizar_excel(config["RUTA_EXCEL"], stock_bsale):
        print("ERROR actualizando Excel. Abortando.")
        return

    productos = leer_produccion(config["RUTA_EXCEL"], stock_bsale)
    if not productos:
        print("ERROR leyendo PRODUCCION. Abortando.")
        return

    html = armar_html(productos)
    enviar_correo(config, html)

    print("=" * 50)
    print("PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 50)

if __name__ == "__main__":
    main()
